# -*- coding: utf-8 -*-
from datetime import timedelta
from django.shortcuts import render, render_to_response, RequestContext, redirect, get_object_or_404
from django.contrib import auth, messages
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_protect
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from models import *
from forms import *


############
# Neutral pages
@login_required
def about(request):
	return render_to_response("about.html", [], context_instance=RequestContext(request))


def access_denied(request):
	return render_to_response("access_denied.html", [], context_instance=RequestContext(request))


def export_documents(request):
	return render_to_response("export_documents.html", locals(), context_instance=RequestContext(request))


@login_required
def home(request):
	return render_to_response("home.html", locals(), context_instance=RequestContext(request))


@csrf_protect
def login(request):
	if request.method == "POST":
		username = request.POST['username']
		password = request.POST['password']
		user = auth.authenticate(username=username, password=password)

		if user is not None and user.is_active:
			auth.login(request, user)
			return redirect('home')
		else:
			messages.error(request, "Nombre de usuario o contraseña incorrecto.")

	return render_to_response("login.html", locals(), context_instance=RequestContext(request))


def logout(request):
	auth.logout(request)
	return redirect(login)


############
# Crops
@login_required
def add_crop(request):
	owner_form = CropOwnerForm(request.POST or None)
	company_form = CompanyCropFrom(request.POST or None)
	crop_form = CropForm(request.POST or None)

	if request.method == "POST":
		if owner_form.is_valid() and company_form.is_valid() and crop_form.is_valid():
			# owner information
			owner = owner_form.cleaned_data['old_owner']
			if owner is None:
				first_name = owner_form.cleaned_data['first_name'].title()
				last_name = owner_form.cleaned_data['last_name'].title()
				rut = owner_form.cleaned_data['rut']
				number_1 = owner_form.cleaned_data['contact_number_1']
				number_2 = owner_form.cleaned_data['contact_number_2']
				email = owner_form.cleaned_data['email']
				position = owner_form.cleaned_data['position'].title()

				# company information
				company = None
				company_member = company_form.cleaned_data['company_member']
				if company_member:
					company = company_form.cleaned_data['excisting_company']
					if company is None:
						company_name = company_form.cleaned_data['comp_name'].title()
						company_rut = company_form.cleaned_data['comp_rut']

						company = CompanyCrop(name=company_name, rut=company_rut)
						company.save()

				owner = CropOwner(first_name=first_name, last_name=last_name, rut=rut, contact_number_1=number_1, contact_number_2=number_2,
					email=email, position=position, company=company)
				owner.save()

			# crop information
			name = crop_form.cleaned_data['name'].title()
			# geographical information
			region = crop_form.cleaned_data['region']
			province = commune = None
			try:
				province = Province.objects.get(name= request.POST['province_trick'])
				commune = Commune.objects.get(name= request.POST['commune'])
			except:
				pass
			address = crop_form.cleaned_data['address']

			# terrain characteristics information
			has = crop_form.cleaned_data['has']
			water = crop_form.cleaned_data['water']
			water_cmnt = crop_form.cleaned_data['water_cmnt']
			soil = crop_form.cleaned_data['soil']
			soil_cmnt = crop_form.cleaned_data['soil_cmnt']
			topo = crop_form.cleaned_data['topography']
			topo_cmnt = crop_form.cleaned_data['topography_cmnt']
			temp = crop_form.cleaned_data['temperatures']
			temp_cmnt = crop_form.cleaned_data['temperatures_cmnt']
			access = crop_form.cleaned_data['access']
			access_cmnt = crop_form.cleaned_data['access_cmnt']

			obs = crop_form.cleaned_data['observations'].strip(' \t\n\r')
			crop = Crop(name=name, region=region, province=province, commune=commune, address=address, has=has,
				water=water, soil=soil, topography=topo, temperatures=temp, access=access,
				water_cmnt=water_cmnt, soil_cmnt=soil_cmnt, topography_cmnt=topo_cmnt, temperatures_cmnt=temp_cmnt, access_cmnt=access_cmnt, observations=obs)
			crop.save()
			crop.crop_owner.add(owner)

			messages.success(request, "Predio agregado exitosamente.")
		else:
			messages.error(request, "Error en el formulario.")

	return render_to_response("crops/add_crop.html", locals(), context_instance=RequestContext(request))


@login_required
def add_plantation(request):
	id = None
	if 'id' in request.GET:
		id = request.GET['id']
		crop = Crop.objects.get(pk=id)

	plantation_form = PlantationForm(request.POST or None)

	if request.method == "POST":
		if plantation_form.is_valid():
			date = plantation_form.cleaned_data['date']
			crop = plantation_form.cleaned_data['crop']
			paddock = plantation_form.cleaned_data['paddock']
			variety = plantation_form.cleaned_data['variety']
			source_lot = plantation_form.cleaned_data['source_lot']
			kg_seeds = plantation_form.cleaned_data['kg_seeds']
			kg_fert = plantation_form.cleaned_data['kg_fert']
			n_mini = plantation_form.cleaned_data['n_mini']
			has = plantation_form.cleaned_data['has']
			control_number = plantation_form.cleaned_data['control_number']
			obs = plantation_form.cleaned_data['obs']

			new_plantation = Plantation(date=date, crop=crop, paddock=paddock, variety=variety, source_lot=source_lot,
				kg_seeds=kg_seeds, kg_fert=kg_fert, n_mini=n_mini, has=has, control_number=control_number, obs=obs)
			new_plantation.save()

			messages.success(request, "Plantacion " + control_number + " registrada con exito.".encode('utf-8'))
			if id is None:
				return redirect('plantation_table')
		else:
			messages.error(request, "Error en el formulario.")

	return render_to_response("crops/add_plantation.html", locals(), context_instance=RequestContext(request))


@login_required
def crops(request):
	return render_to_response("crops/crops.html", locals(), context_instance=RequestContext(request))


@login_required
def crop_info(request):
	if not 'id' in request.GET:
		return redirect('crop_table')
	id = request.GET['id']

	try:
		crop = Crop.objects.get(pk=id)
	except:
		return redirect('crop_table')
	owner = crop.crop_owner.first()
	comp = owner.company
	plantations = Plantation.objects.filter(crop=crop).order_by('-date')[:6]

	crop_form = CropForm(request.POST or None)
	owner_form = CropOwnerForm(request.POST or None)
	company_form = CompanyCropFrom(request.POST or None)

	# Delete crop
	if 'delete' in request.POST:
		crop.delete()
		return redirect('crop_table')

	# Edit section
	if request.method == "POST":
		if owner_form.is_valid() and crop_form.is_valid():
			owner.contact_number_1 = owner_form.cleaned_data['contact_number_1']
			owner.contact_number_2 = owner_form.cleaned_data['contact_number_2']
			owner.email = owner_form.cleaned_data['email']
			owner.save()

			crop.has = crop_form.cleaned_data['has']
			crop.water = crop_form.cleaned_data['water']
			crop.water_cmnt= crop_form.cleaned_data['water_cmnt']
			crop.soil = crop_form.cleaned_data['soil']
			crop.soil_cmnt = crop_form.cleaned_data['soil_cmnt']
			crop.topography = crop_form.cleaned_data['topography']
			crop.topography_cmnt = crop_form.cleaned_data['topography_cmnt']
			crop.temperatures = crop_form.cleaned_data['temperatures']
			crop.temperatures_cmnt = crop_form.cleaned_data['temperatures_cmnt']
			crop.access = crop_form.cleaned_data['access']
			crop.access_cmnt = crop_form.cleaned_data['access_cmnt']

			crop.region = crop_form.cleaned_data['region']
			province = commune = None
			try:
				province = Province.objects.get(name= request.POST['province_trick'])
				commune = Commune.objects.get(name= request.POST['commune'])
			except:
				pass
			crop.province = province
			crop.commune = commune
			crop.address = crop_form.cleaned_data['address']
			crop.observations = crop_form.cleaned_data['observations']
			crop.save()

			messages.success(request, "Edición guardada con éxito.")
		else: 
			messages.error(request, "Error en el formulario de edición.")

	crop.score = assign_score(crop)
	return render_to_response("crops/crop_info.html", locals(), context_instance=RequestContext(request))


@login_required
def crop_map(request):
	crops = Crop.objects.all()
	for crop in crops:
		crop.score = assign_score(crop)
	return render_to_response("crops/crop_map.html", locals(), context_instance=RequestContext(request))


@login_required
def crop_table(request):
	crops = Crop.objects.all()
	for crop in crops:
		crop.score = assign_score(crop)
	return render_to_response("crops/crop_table.html", locals(), context_instance=RequestContext(request))


def assign_score(crop):
	"""Returns the score of a crop, to later be translated into ☆'stars."""
	score = {"water": 40, "soil": 20, "topo": 15, "weather": 15, "access": 10} # weight for each property
	crop_score = 0
	if crop.water:
		crop_score += score["water"]
	if crop.soil:
		crop_score += score["soil"]
	if crop.topography:
		crop_score += score["topo"]
	if crop.temperatures:
		crop_score += score["weather"]
	if crop.access :
		crop_score += score["access"]
	return crop_score


@login_required
def export_crops_xlsx(request):
	# todo: add security. filter by request.user.groups

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename=crops.xlsx'

	wb = Workbook()
	ws = wb.active
	ws.title = "Crops"

	# Crops
	row_num = 0
	columns = ["Predio", "Dueño", "Has", "Región", "Provincia", "Comuna", "Dirección",
		"Agua", "Obs. agua", "Suelo", "Obs. suelo", "Topo", "Obs. topo", "Clima", "Obs. clima", "Acceso", "Obs. acceso", 
		"Observaciones"
	]

	for col_num in xrange(len(columns)):
		c = ws.cell(row=1, column=col_num+1)
		c.value = columns[col_num]
		c.font = Font(bold=True)
		c.alignment = Alignment(horizontal='center')

	queryset = Crop.objects.all()

	for crop in queryset:
		row_num += 1
		row = [crop.name, str(crop.crop_owner.first()), crop.has, str(crop.region), str(crop.province), str(crop.commune), crop.address,
			crop.water, crop.water_cmnt, crop.soil, crop.soil_cmnt, crop.topography, crop.topography_cmnt, crop.temperatures, crop.temperatures_cmnt,
			crop.access, crop.access_cmnt, crop.observations
		]
		for col_num in xrange(len(row)):
			c = ws.cell(row=row_num+1, column=col_num+1)
			c.value = row[col_num]
			if row[col_num] == True:
				c.font = Font(color='006100')
			if row[col_num] == False:
				c.font = Font(color='9C0006')
			c.alignment = Alignment(horizontal='center')

	# Owners
	ws = wb.create_sheet()
	ws.title = "Props"

	row_num = 0
	columns = ["Nombre", "Apellido", "Compañía", "Rut", "Número 1", "Número 2", "Email"]

	for col_num in xrange(len(columns)):
		c = ws.cell(row=1, column=col_num+1)
		c.value = columns[col_num]
		c.font = Font(bold=True)
		c.alignment = Alignment(horizontal='center')

	queryset = CropOwner.objects.all()

	for owner in queryset: 
		row_num += 1
		row = [owner.first_name, owner.last_name, str(owner.company), owner.rut, owner.contact_number_1, owner.contact_number_2, owner.email]
		for col_num in xrange(len(row)):
			c = ws.cell(row=row_num+1, column=col_num+1)
			c.value = row[col_num]
			c.alignment = Alignment(horizontal='center')

	# Plantations
	ws = wb.create_sheet()
	ws.title = "Plantaciones"

	row_num = 0
	columns = ["Fecha", "Campo", "Potrero", "Variedad", "Lote Origen", "Kg. Semillas", "Kg. Fert.", "N° Mini", "Has", "N° Control Final SAG", "Observaciones"]

	for col_num in xrange(len(columns)):
		c = ws.cell(row=1, column=col_num+1)
		c.value = columns[col_num]
		c.font = Font(bold=True)
		c.alignment = Alignment(horizontal='center')

	queryset = Plantation.objects.all()

	for plant in queryset: 
		row_num += 1
		row = [plant.date, str(plant.crop), plant.paddock, str(plant.variety), plant.source_lot, plant.kg_seeds, plant.kg_fert, 
			plant.n_mini, plant.has, plant.control_number, plant.obs]
		for col_num in xrange(len(row)):
			c = ws.cell(row=row_num+1, column=col_num+1)
			c.value = row[col_num]
			c.alignment = Alignment(horizontal='center')

	# game set & match.
	wb.save(response)
	return response


@login_required
def photo_library(request):
	if not 'id' in request.GET:
		return redirect('crop_table')
	id = request.GET['id']
	crop = Crop.objects.get(pk=id)

	if request.method == "POST":
		if 'image' in request.FILES:
			img = request.FILES['image']
			new_image = CropImage(crop=crop, image=img)
			new_image.save()
			messages.success(request, "Imagen guardada con éxito")

	images = CropImage.objects.filter(crop=crop)

	return render_to_response("crops/photo_library.html", locals(), context_instance=RequestContext(request))


@login_required
def plantation_info(request):
	if not 'cn' in request.GET:
		return redirect('plantation_table')
	control_number = request.GET['cn']
	control_number = control_number.replace('+', ' ')
	plantation = Plantation.objects.get(control_number=control_number)

	plantation_form = PlantationForm(request.POST or None)
	# plantation_form.date = plantation.date 

	# Edit
	if request.method == 'POST':
		if plantation_form.is_valid():
			#date = plantation_form.cleaned_data['date']  # fix this, edit wont work
			plantation.crop = plantation_form.cleaned_data['crop']
			plantation.paddock = plantation_form.cleaned_data['paddock']
			plantation.variety = plantation_form.cleaned_data['variety']
			plantation.source_lot = plantation_form.cleaned_data['source_lot']
			plantation.kg_seeds = plantation_form.cleaned_data['kg_seeds']
			plantation.kg_fert = plantation_form.cleaned_data['kg_fert']
			plantation.n_mini = plantation_form.cleaned_data['n_mini']
			plantation.has = plantation_form.cleaned_data['has']
			plantation.control_number = plantation_form.cleaned_data['control_number']
			plantation.obs = plantation_form.cleaned_data['obs']

			plantation.save()
			message.success(request, "Edición realizada con éxito")
		else: 
			messages.error(request, "Error en el formulario")

	return render_to_response("crops/plantation_info.html", locals(), context_instance=RequestContext(request))


@login_required
def plantation_table(request):
	plantations = Plantation.objects.all()
	return render_to_response("crops/plantation_table.html", locals(), context_instance=RequestContext(request))


############
# Markets
@login_required
def add_market(request):
	client_form = ClientForm(request.POST or None)
	client_type_form = ClientTypeForm(request.POST or None)
	geographical_form = GeoMarkerForm(request.POST or None)
	company_form = CompanyMarketForm(request.POST or None)

	if request.method == 'POST':
		if client_form.is_valid() and client_type_form.is_valid() and geographical_form.is_valid() and company_form.is_valid():
			# client information
			type_of_client = client_type_form.cleaned_data['type_of_client']
			first_name = client_form.cleaned_data['first_name'].title()
			last_name = client_form.cleaned_data['last_name'].title()
			rut = client_form.cleaned_data['rut']
			number_1 = client_form.cleaned_data['contact_number_1']
			number_2 = client_form.cleaned_data['contact_number_2']
			email = client_form.cleaned_data['email']
			position = client_form.cleaned_data['position'].title()
			obs = client_form.cleaned_data['observations'].strip(' \t\n\r')

			# company information
			company = company_form.cleaned_data['excisting_company']
			if company is None:
				company_name = company_form.cleaned_data['name'].title()
				company_rut = company_form.cleaned_data['rut']
				if company_name != "":
					company = CompanyMarket(name=company_name, rut=company_rut)
					company.save()

			new_client = Client(type_of_client=type_of_client, first_name=first_name, last_name=last_name, rut=rut, contact_number_1=number_1, 
				contact_number_2=number_2, email=email, position=position, observations=obs, company=company)
			new_client.save()

			# geographical information
			region = geographical_form.cleaned_data['region']
			province = None
			commune = None
			try:
				province = Province.objects.get(name= request.POST['province_trick'])
				commune = Commune.objects.get(name= request.POST['commune'])
			except:
				pass
			address = geographical_form.cleaned_data['address']

			new_geomarker = GeoMarker(client=new_client, region=region, province=province, commune=commune, address=address)
			new_geomarker.save()

			# all done -- success
			messages.success(request, "Cliente agregado exitosamente.")
		else:
			messages.error(request, "Error en el formulario.")

	return render_to_response("markets/add_market.html", locals(), context_instance=RequestContext(request))


@login_required
def add_sale(request):
	sale_form = TransactionForm(request.POST or None)
	sale_detail_formset = SaleDetailFormSet(request.POST or None, prefix="form")

	if request.method == 'GET':
		client_id = request.GET['id']
		client = Client.objects.get(pk=client_id)

	if request.method == 'POST':
		client_id = request.GET['id']
		if sale_form.is_valid() and sale_detail_formset.is_valid():
			user = request.user
			date = sale_form.cleaned_data['date']
			type_of_transaction = sale_form.cleaned_data['type_of_transaction']
			obs = sale_form.cleaned_data['observations'].strip(' \t\n\r')

			new_sale = Sale(user=user, client=Client.objects.get(pk=client_id), type_of_transaction=type_of_transaction, date=date, observations=obs)
			new_sale.save()

			for sale_detail in sale_detail_formset:
				price = sale_detail.cleaned_data['price']
				volume = sale_detail.cleaned_data['volume']
				variety = sale_detail.cleaned_data['variety']
				certification = sale_detail.cleaned_data['certification']

				sdetail = SaleDetail(sale=new_sale, price=price, volume=volume, variety=variety, certificate=certification)
				sdetail.save()

			messages.success(request, 'Venta de agregada con éxtio.')
		else:
			messages.error(request, 'Error en el formulario.')

	return render_to_response("markets/add_sale.html", locals(), context_instance=RequestContext(request))


@login_required
def export_markets_xlsx(request):
	# todo: add security. filter by request.user.groups

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename=markets.xlsx'

	# Clients
	wb = Workbook()
	ws_client = wb.active
	ws_potential = wb.create_sheet()
	ws_client.title = "Clientes Actuales"
	ws_potential.title = "Clientes Potenciales"

	row_num_0 = row_num_1 = 0
	columns = ["Cliente", "Compañía", "Número 1", "Número 2", "Email", "Región", "Provincia", "Comuna", "Dirección", "Observaciones"]

	for col_num in xrange(len(columns)):
		c0 = ws_client.cell(row=1, column=col_num+1)
		c1 = ws_potential.cell(row=1, column=col_num+1)
		c0.value = c1.value = columns[col_num]
		c0.font = c1.font = Font(bold=True)

	queryset = GeoMarker.objects.all()

	for obj in queryset:
		if obj.client.type_of_client.type == "Actual": 
			ws = ws_client
			row_num_0 += 1
			ws.row_num = row_num_0 
		else:
			ws = ws_potential
			row_num_1 += 1
			ws.row_num = row_num_1

		row = [str(obj.client), "Compañía", obj.client.contact_number_1, obj.client.contact_number_2, obj.client.email, 
			str(obj.region), str(obj.province), str(obj.commune), obj.address, obj.client.observations]
		for col_num in xrange(len(row)):
			c = ws.cell(row=ws.row_num+1, column=col_num+1)
			c.value = row[col_num]

	# Transacciones
	ws_reserves = wb.create_sheet()
	ws_sales = wb.create_sheet()
	ws_reserves.title = "Reservas"
	ws_sales.title = "Ventas"

	row_num_0 = row_num_1 = 0
	columns = ["Fecha", "Cliente", "Monto", "Volumen", "Variedades"]

	for col_num in xrange(len(columns)):
		c0 = ws_sales.cell(row=1, column=col_num+1)
		c1 = ws_reserves.cell(row=1, column=col_num+1)
		c0.value = c1.value = columns[col_num]
		c0.font = c1.font = Font(bold=True)

	queryset = Sale.objects.all()

	for obj in queryset:
		if obj.type_of_transaction.type == "Reserva": # reserve
			ws = ws_reserves
			row_num_0 += 1
			ws.row_num = row_num_0
		else: # sale
			ws = ws_sales
			row_num_1 += 1
			ws.row_num = row_num_1

		row = [obj.date, str(obj.client), '$'+str(obj.get_price()), obj.get_volume(), obj.get_varieties()]
		for col_num in xrange(len(row)):
			c = ws.cell(row=ws.row_num+1, column=col_num+1)
			c.value = row[col_num]

	# all done -- download the .xlsx file
	wb.save(response)
	return response


@login_required
def markets(request):
	return render_to_response("markets/markets.html", [], context_instance=RequestContext(request))


@login_required
def market_company(request):
	if not 'id' in request.GET:
		return redirect('market_table')
	id = request.GET['id']

	company = CompanyMarket.objects.get(pk=id)
	clients = Client.objects.filter(company=company)

	# Calculating data
	n = total_volume = total_price = 0
	varieties = {}
	now = datetime.datetime.now()
	for client in clients:
		if client.type_of_client.type == "Actual":
			sales = Sale.objects.filter(client=client, type_of_transaction=2, date__range=(now - timedelta(3*365), now + timedelta(365)))
			for sale in sales:
				n += 1
				for sdetail in sale.saledetail_set.all():
					total_volume += sdetail.volume
					total_price += sdetail.price
					if sdetail.variety not in varieties:
						varieties[sdetail.variety] = sdetail.volume
					else:
						varieties[sdetail.variety] += sdetail.volume
	avg_volume = avg_price = 0
	if n != 0:
		for var in varieties:
			varieties[var] = "{0:.2f}".format(100.0 * varieties[var] / total_volume)
		avg_volume = total_volume / n
		avg_price = total_price / n

	# Edit
	company_form = CompanyMarketForm(request.POST or None)
	if request.method == "POST":
		if company_form.is_valid():
			company.name = company_form.cleaned_data['name'].title()
			company.rut = company_form.cleaned_data['rut']
			company.save()
			messages.success(request, 'Edición guardada con éxito')
		else: 
			messages.error(request, 'Error en la edición')

	return render_to_response("markets/market_company.html", locals(), context_instance=RequestContext(request))


@login_required
def market_info(request):
	if not 'id' in request.GET:
		return redirect('market_table')
	id = request.GET['id']

	client_form = ClientForm(request.POST or None)
	geographical_form = GeoMarkerForm(request.POST or None)
	company_form = CompanyMarketForm(request.POST or None)
	client = Client.objects.get(pk = id)
	geo_info = GeoMarker.objects.get(client = id) # change to filter. this should allow multiple locations.

	# Delete client
	if 'delete' in request.POST:
		type_of_client = client.type_of_client.type
		sales = client.sale_set.all()
		for sale in sales:
			sale.saledetail_set.all().delete()
		sales.delete()
		geo_info.delete()
		client.delete()
		messages.success(request, "Cliente removido de la base de datos.")
		if type_of_client == "Actual":
			return redirect("market_table")
		else: 
			return redirect("market_table_potential")

	# Calculate variaty distribution
	if client.type_of_client.type == "Actual":
		now = datetime.datetime.now()
		sales = Sale.objects.filter(client=client, type_of_transaction=2, date__range=(now - timedelta(3*365), now + timedelta(365)))
		total_volume = 0
		n = len(sales)
		if n != 0:
			total_price = 0
			varieties = {}
			for sale in sales:
				for s in sale.saledetail_set.all():
					total_price += s.price
					total_volume += s.volume
					if s.variety not in varieties:
						varieties[s.variety] = s.volume
					else:
						varieties[s.variety] += s.volume
			for var in varieties:
				varieties[var] = "{0:.2f}".format(100.0 * varieties[var] / total_volume)
			avg_price = total_price / n
			avg_volume = total_volume / n
		else:
			avg_price = avg_volume = 0

	# Edit section
	if request.method == "POST":
		if client_form.is_valid() and geographical_form.is_valid():
			# Upgrades client from potential to actual
			if 'upgrade' in request.POST:
				client.type_of_client = TypeOfClient.objects.get(type="Actual")
			client.first_name = client_form.cleaned_data['first_name'].title()
			client.last_name = client_form.cleaned_data['last_name'].title()
			client.rut = client_form.cleaned_data['rut']
			client.contact_number_1 = client_form.cleaned_data['contact_number_1']
			client.contact_number_2 = client_form.cleaned_data['contact_number_2']
			client.email = client_form.cleaned_data['email']
			client.position = client_form.cleaned_data['position'].title()
			client.observations = client_form.cleaned_data['observations'].strip(' \t\n\r')
			client.save()

			geo_info.region = geographical_form.cleaned_data['region']
			province = commune = None
			try:
				province = Province.objects.get(name= request.POST['province_trick'])
				commune = Commune.objects.get(name= request.POST['commune'])
			except:
				pass
			geo_info.province = province
			geo_info.commune = commune
			geo_info.address = geographical_form.cleaned_data['address']
			geo_info.save()

			messages.success(request, 'Edicion guardada con éxito.')
		else: 
			messages.error(request, 'Error en la edición.')

	return render_to_response("markets/market_info.html", locals(), context_instance=RequestContext(request))


@login_required
def market_map(request):
	geomarkers = GeoMarker.objects.all()
	return render_to_response("markets/market_map.html", locals(), context_instance=RequestContext(request))


@login_required
def market_table(request):
	clients = Client.objects.filter(type_of_client=TypeOfClient.objects.get(type="Actual"))
	total_sale_volume = 0
	for client in clients:
		now = datetime.datetime.now()
		sales_last_3_years = client.sale_set.filter(type_of_transaction=2, date__range=(now - timedelta(3*365), now + timedelta(365)))
		client.volume = 0
		for sale in sales_last_3_years:
			sale_volume = sale.get_volume()
			total_sale_volume += sale_volume
			client.volume += sale_volume

	for client in clients:
		client.size = translate_size(total_sale_volume, client.volume)

	return render_to_response("markets/market_table.html", locals(), context_instance=RequestContext(request))


# def calculate_client_size(client):
# 	# returns the size of a client; {'xs', 's', 'm', 'l', 'xl'}
# 	return 2


def translate_size(total_volume, client_volume):
	if total_volume == 0:
		return "0"
	ratio = client_volume / (total_volume * 1.0)
	if ratio == 0:
		return "0"
	if ratio < 0.2:
		return "XS" 
	if 0.2 < ratio and ratio < 0.4:
		return "S"
	if 0.4 < ratio and ratio < 0.6:
		return "M"
	if 0.6 < ratio and ratio < 0.8:
		return "L"
	if 0.8 < ratio:
		return "XL"


@login_required
def market_table_potential(request):
	potential_table = True
	clients = Client.objects.filter(type_of_client=TypeOfClient.objects.get(type="Potencial"))
	return render_to_response("markets/market_table.html", locals(), context_instance=RequestContext(request))


@login_required
def sales_detail(request):
	if request.method == 'GET':
		sale_id = request.GET['id']
		sale = Sale.objects.get(pk=sale_id)
		sale_details = sale.saledetail_set.all()
		previous_id = sale.client.pk
	return render_to_response("markets/sales_detail.html", locals(), context_instance=RequestContext(request))


@login_required
def sales_history(request):
	if request.method == 'GET':
		id = request.GET['id']
		client = Client.objects.get(pk = id)
		sales = Sale.objects.filter(client = id)

		for sale in sales:
			sale.total_volume = 0
			sale.varieties = {}
			for sd in sale.saledetail_set.all():
				sale.total_volume += sd.volume
				if sd.variety not in sale.varieties:
					sale.varieties[sd.variety] = sd.volume
				else:
					sale.varieties[sd.variety] += sd.volume
			
			for var in sale.varieties:
				sale.varieties[var] = "{0:.2f}".format(100.0 * sale.varieties[var] / sale.total_volume)

	return render_to_response("markets/sales_history.html", locals(), context_instance=RequestContext(request))


############
# Related Contacts 
@login_required
def add_related(request):
	related_form = ContactForm(request.POST or None)

	if request.method == 'POST':
		if related_form.is_valid():
			first_name = related_form.cleaned_data['first_name'].title()
			last_name = related_form.cleaned_data['last_name'].title()
			rut = related_form.cleaned_data['rut']
			number_1 = related_form.cleaned_data['contact_number_1']
			number_2 = related_form.cleaned_data['contact_number_2']
			email = related_form.cleaned_data['email']
			obs = related_form.cleaned_data['observations'].strip(' \t\n\r')
			area = related_form.cleaned_data['related_area']

			new_contact = Related(area=area, first_name=first_name, last_name=last_name, rut=rut, contact_number_1=number_1, contact_number_2=number_2,
				email=email, observations=obs)
			new_contact.save()

			messages.success(request, 'Contacto agregado con éxtio.')
		else:
			messages.error(request, 'Error en el formulario.')

	return render_to_response("related/add_related.html", locals(), context_instance=RequestContext(request))


@login_required
def related(request):
	return render_to_response("related/related.html", [], context_instance=RequestContext(request))


@login_required
def related_info(request):
	if not 'id' in request.GET:
		return redirect('related_table')
	id = request.GET['id']

	contact_form = ContactForm(request.POST or None)
	contact = Related.objects.get(pk=id)

	# Edit section
	if request.method == "POST":
		if contact_form.is_valid():
			contact.first_name = contact_form.cleaned_data['first_name'].title()
			contact.last_name = contact_form.cleaned_data['last_name'].title()
			contact.rut = contact_form.cleaned_data['rut']
			contact.contact_number_1 = contact_form.cleaned_data['contact_number_1']
			contact.contact_number_2 = contact_form.cleaned_data['contact_number_2']			
			contact.email = contact_form.cleaned_data['email']
			contact.observations = contact_form.cleaned_data['observations'].strip(' \t\n\r')
			contact.save()

			messages.success(request, 'Edición guardada con éxtito')
		else:
			messages.error(request, 'Error en la edición')

	return render_to_response("related/related_info.html", locals(), context_instance=RequestContext(request))


@login_required
def related_table(request):
	contacts = Related.objects.all()
	return render_to_response("related/related_table.html", locals(), context_instance=RequestContext(request))

